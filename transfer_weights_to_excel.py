from openpyxl import Workbook

#Just creating the list variables. No need to look at this part.
one_year = [('SNX', 0.00474), ('SUSHI', 0.00132), ('UMA', 0.00573), ('XLM', 0.00601), ('XRP', 0.00537), ('XTZ', 0.0), ('YFI', 0.14032), ('ZRX', 0.00151), ('LINK', 0.00605), ('LRC', 0.00299), ('LTC', 0.00423), ('MATIC', 0.19967), ('MIR', 0.00237), ('NKN', 0.01953), ('OMG', 0.00311), ('OXT', 0.0008), ('BTC', 0.00563), ('CELO', 0.00426), ('COMP', 0.0), ('CRV', 0.0), ('DOGE', 0.48095), ('ENJ', 0.01257), ('EOS', 0.00089), ('ETH', 0.01807), ('FIL', 0.00761), ('AAVE', 0.0), ('ADA', 0.04444), ('ALGO', 0.006), ('BAND', 0.00798), ('BAT', 0.00186), ('BCH', 0.00202), ('BNT', 0.00389)]
six_months = [('SNX', 0.0), ('SUSHI', 0.0), ('UMA', 0.0), ('XLM', 0.0), ('XRP', 0.0), ('XTZ', 0.0), ('YFI', 0.0), ('ZRX', 0.0), ('LINK', 0.0), ('LRC', 0.0), ('LTC', 0.0), ('MATIC', 0.26232), ('MIR', 0.00386), ('NKN', 0.0), ('OMG', 0.0), ('OXT', 0.0), ('BTC', 0.0), ('CELO', 0.0), ('COMP', 0.0), ('CRV', 0.0), ('DOGE', 0.73382), ('ENJ', 0.0), ('EOS', 0.0), ('ETH', 0.0), ('FIL', 0.0), ('AAVE', 0.0), ('ADA', 0.0), ('ALGO', 0.0), ('BAND', 0.0), ('BAT', 0.0), ('BCH', 0.0), ('BNT', 0.0)]
three_months = [('SNX', 0.0), ('SUSHI', 0.0), ('UMA', 0.0), ('XLM', 0.0), ('XRP', 0.0), ('XTZ', 0.0), ('YFI', 0.0), ('ZRX', 0.0), ('LINK', 0.0), ('LRC', 0.0), ('LTC', 0.0), ('MATIC', 0.25981), ('MIR', 0.00415), ('NKN', 0.0), ('OMG', 0.0), ('OXT', 0.0), ('BTC', 0.0), ('CELO', 0.0), ('COMP', 0.0), ('CRV', 0.0), ('DOGE', 0.73603), ('ENJ', 0.0), ('EOS', 0.0), ('ETH', 0.0), ('FIL', 0.0), ('AAVE', 0.0), ('ADA', 0.0), ('ALGO', 0.0), ('BAND', 0.0), ('BAT', 0.0), ('BCH', 0.0), ('BNT', 0.0)]
one_month = [('SNX', 0.0), ('SUSHI', 0.0), ('UMA', 0.0), ('XLM', 0.0), ('XRP', 0.0), ('XTZ', 0.0), ('YFI', 0.0), ('ZRX', 0.0), ('LINK', 0.0), ('LRC', 0.0), ('LTC', 0.0), ('MATIC', 0.26841), ('MIR', 0.00321), ('NKN', 0.0), ('OMG', 0.0), ('OXT', 0.0), ('BTC', 0.0), ('CELO', 0.0), ('COMP', 0.0), ('CRV', 0.0), ('DOGE', 0.72838), ('ENJ', 0.0), ('EOS', 0.0), ('ETH', 0.0), ('FIL', 0.0), ('AAVE', 0.0), ('ADA', 0.0), ('ALGO', 0.0), ('BAND', 0.0), ('BAT', 0.0), ('BCH', 0.0), ('BNT', 0.0)]
fifteen_days = [('SNX', 0.0), ('SUSHI', 0.0), ('UMA', 0.0), ('XLM', 0.0), ('XRP', 0.0), ('XTZ', 0.0), ('YFI', 0.0), ('ZRX', 0.0), ('LINK', 0.0), ('LRC', 0.0), ('LTC', 0.0), ('MATIC', 0.0), ('MIR', 0.13798), ('NKN', 0.0), ('OMG', 0.0), ('OXT', 0.0), ('BTC', 0.0), ('CELO', 0.86202), ('COMP', 0.0), ('CRV', 0.0), ('DOGE', 0.0), ('ENJ', 0.0), ('EOS', 0.0), ('ETH', 0.0), ('FIL', 0.0), ('AAVE', 0.0), ('ADA', 0.0), ('ALGO', 0.0), ('BAND', 0.0), ('BAT', 0.0), ('BCH', 0.0), ('BNT', 0.0)]
#we can automate this part later if it's required.

wb = Workbook()
ws = wb.active
count = 0
for tp in one_year:
    count += 1
    ws[f"A{count}"] = tp[0]
    ws[f"B{count}"] = tp[1]


count = 0
for tp in three_months:
    count += 1
    ws[f"D{count}"] = tp[1]


count = 0
for tp in six_months:
    count += 1
    ws[f"C{count}"] = tp[1]

count = 0
for tp in one_month:
    count += 1
    ws[f"E{count}"] = tp[1]

count = 0
for tp in fifteen_days:
    count += 1
    ws[f"F{count}"] = tp[1]

wb.save(filename="invest_ratios.xlsx")

#sorry for the redundant code but i don't want to waste time writing efficient code for something
#that we may not use again.

